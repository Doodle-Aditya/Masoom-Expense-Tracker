import os
from openpyxl import Workbook, load_workbook

DB_PATH = os.path.join("output", "records.xlsx")


def ensure_excel_exists():
    if not os.path.exists(DB_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Ref No", "Date", "College", "Class", "Student Name",
            "College Fees", "Masoom 75%", "Student 25%", "Payable",
            "Rs In Words", "Donor", "Cheque Issue", "Account Holder",
            "Bank Name", "Account Number", "IFSC", "Cheque Number", "Prepared By"
        ])
        wb.save(DB_PATH)


def save_to_excel(data: dict):
    ensure_excel_exists()

    wb = load_workbook(DB_PATH)
    ws = wb.active

    ws.append([
        data["ref_no"], data["date"], data["college_name"], data["class"],
        data["student_name"], data["college_fees"], data["masoom_contribution"],
        data["student_contribution"], data["payable"], data["rs_in_words"],
        data["donor_name"], data["cheque_issue_name"], data["account_holder"],
        data["bank_name"], data["account_number"], data["ifsc"],
        data["cheque_number"], data["prepared_by"]
    ])

    wb.save(DB_PATH)


def next_ref_no():
    ensure_excel_exists()

    wb = load_workbook(DB_PATH)
    ws = wb.active
    last_row = ws.max_row

    if last_row == 1:
        return "M-0001/24-25/LT"

    last_ref = ws.cell(row=last_row, column=1).value
    num = int(last_ref.split("-")[1].split("/")[0])

    return f"M-{num+1:04d}/24-25/LT"


def load_donors_from_pdf():
    return ["NTT", "Deautche Bank (DB) All School"," NSTP,	ELC	& College" "Other"]
