import os
from openpyxl import load_workbook
import win32com.client as win32

TEMPLATE = os.path.join("excel_template", "bill_template.xlsx")


def fill_excel_template(data, filled_path):
    wb = load_workbook(TEMPLATE)
    ws = wb.active

    # Map Excel cells
    ws["B3"] = data["college_name"]
    ws["B4"] = data["class"]
    ws["E3"] = data["ref_no"]
    ws["E4"] = data["date"]

    ws["B6"] = data["student_name"]
    ws["B10"] = data["college_fees"]
    ws["E10"] = data["payable"]
    ws["B12"] = data["rs_in_words"]
    ws["B14"] = data["donor_name"]

    ws["B20"] = data["cheque_issue_name"]
    ws["B21"] = data["account_holder"]
    ws["B22"] = data["bank_name"]
    ws["B23"] = data["account_number"]
    ws["B24"] = data["ifsc"]
    ws["B25"] = data["cheque_number"]

    ws["E30"] = data["prepared_by"]

    wb.save(filled_path)


def excel_to_pdf(excel_path, pdf_path):
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False

    wb = excel.Workbooks.Open(os.path.abspath(excel_path))
    wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))

    wb.Close()
    excel.Quit()
